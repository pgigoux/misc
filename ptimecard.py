import sys
import re
from typing import Any
from argparse import ArgumentParser, SUPPRESS, Namespace
import pandas as pd
from math import modf
from numpy import int64, float64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Project file name
PROJECT_FILE = 'projects.txt'

# Column names
COL_PROJECT_PATH = 'Project Path'
COL_PROJECT_CODE = 'Project Code'
COL_BILLABLE = 'Billable'
COL_BILLABLE_AMOUNT = 'Billable Amount'
COL_TOTAL_HOURS = 'Total Hours'  # original name of the column, renamed at run-time as COL_TOTALS
COL_TOTALS = 'Totals'
COL_ERRORS = 'Errors'
COL_FORMULAS = 'Formulas'

# Row names
ROW_TOTALS = 'Totals'
ROW_FORMULAS = 'Formulas'

# Column types
UNNEEDED_COLUMNS = [COL_PROJECT_CODE, COL_BILLABLE, COL_BILLABLE_AMOUNT]
NON_INPUT_DATA_COLUMNS = [COL_PROJECT_PATH, COL_TOTAL_HOURS, COL_TOTALS, COL_ERRORS]

# Font name used to format cells
FONT_NAME = 'Liberation Sans'


def dump_data(df: pd.DataFrame, label='', print_types=False):
    """
    Print the contents of a pandas DataFrame on the screen.
    Used for testing purposes.
    :param df: pandas data frame
    :param label: label to print in output header
    :param print_types: print data types along with values
    """
    assert (isinstance(df, pd.DataFrame))
    aux = '-- ' + label + ' ' if label else ''
    delimiter = aux + '-' * 100
    print(delimiter)
    print('+ ', end='')
    for col in df.keys():
        print('[' + str(col) + ']', end=' ')
    print('\n')
    for row in list(df.index.values):
        for col in df.keys():
            value = df.at[row, col]
            if print_types:
                print('[' + str(value) + ' ' + str(type(value)) + ']', end=' ')
            else:
                print('[' + str(value) + ']', end=' ')
        print()
    print()


def closest_number(x: Any):
    """
    Return the closest real to a given number whose fractional part is either 0, 0.25, 0.5 or 0.75.
    :param x: number to approximate
    :return: tuple with closest number and difference between the two
    """
    xf, xi = modf(x)
    if xf > 0.875:
        xi += 1
        xf = 0
    elif xf > 0.675:
        xf = 0.75
    elif xf > 0.375:
        xf = 0.5
    elif xf > 0.125:
        xf = 0.25
    else:
        xf = 0
    xt = xi + xf
    dx = x - xt if abs(x - xt) > 1.0E-4 else 0
    return xt, dx


def read_project_list(file_name: str) -> list:
    """
    Read the contents of the file with the list of valid projects.
    :param file_name: project list file name
    :return: list of projects
    """
    output_list = []
    f = open(file_name, 'r')
    for line in f:
        line = line.strip()
        if not re.search('^#', line) or len(line) == 0:
            # print(line)
            output_list.append(line)
    # print(lst)
    return output_list


def check_projects(df: pd.DataFrame, project_list: list) -> bool:
    """
    Make sure that all the projects in the data frame are named in the master project list.
    :param df: data frame
    :param project_list: list of valid projects
    :return: False if a project is not in the master list, False otherwise
    """
    assert (isinstance(df, pd.DataFrame))
    ok = True
    for _, row in df.iterrows():
        project_name = row[COL_PROJECT_PATH]
        if project_name not in project_list and project_name != ROW_TOTALS:
            print(project_name + ' not in project list')
            ok = False
    return ok


def read_excel_file(file_name: str) -> pd.DataFrame:
    """
    Read Excel spreadsheet into a Pandas data frame
    :param file_name: spreadsheet file name
    :return: pandas data frame
    :rtype: pd.DataFrame
    """
    df = pd.read_excel(file_name)

    # Drop unused columns
    df_out = df.drop(columns=UNNEEDED_COLUMNS)

    # Look for the row containing the column totals
    lst = df_out.index[df_out[COL_PROJECT_PATH] == ROW_TOTALS].tolist()
    if len(lst) == 1:
        index = lst[0]
    else:
        raise (ValueError, 'No totals')

    # Only return rows up to the totals
    return df_out[0:index + 1]


def convert_values(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert hh:mm data to decimal, making sure that times bigger than 24 hours are handled correctly.
    Convert all numeric cells to int64 or float64 data.
    :param df: data frame
    :return: converted data frame
    """
    assert (isinstance(df, pd.DataFrame))

    # Make a copy of the data frame so the original is left untouched
    df_out = df.copy(deep=True)

    for row_index, _ in df_out.iterrows():
        for col_name in df_out.keys():
            value = df_out.at[row_index, col_name]
            try:
                h, m = map(int, str(value).split(':'))
                result = h + m / 60.0
                df_out.at[row_index, col_name] = result
            except ValueError:
                pass

    # Then convert all columns (except for the project name/path) to the same data type
    # (should be numpy.float64)
    for col_name in df_out.keys():
        if col_name != COL_PROJECT_PATH:
            df_out[col_name] = pd.to_numeric(df_out[col_name])

    return df_out


def rename_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Rename date of the week, which originally are of the form 'dow yyyy-mm-dd', to 'dow day'.
    Also rename the column 'Total Hours' to 'Totals'
    :param df: input data frame
    :return: converted data frame
    """
    assert (isinstance(df, pd.DataFrame))
    d = {COL_TOTAL_HOURS: COL_TOTALS}
    for col in df.keys():
        if col not in NON_INPUT_DATA_COLUMNS:
            day_of_week, rest = col.split()
            day = rest.split('/')[0]
            d[col] = day_of_week + ' ' + day
    df_out = df.rename(columns=d)
    return df_out


def round_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Round time data to the nearest 15 minute boundary (x.0 x.25, xx.5, x.75).
    The rounding errors are recorded into the (new) 'Errors' column.
    The row and column totals become inconsistent after this point.
    :param df: input data frame
    :return: converted data frame
    """
    assert (isinstance(df, pd.DataFrame))

    # Make a copy of the data frame so the original is left untouched
    df_out = df.copy(deep=True)

    # First round out every data cell to the nearest number.
    # Keep track of the errors for the row in the errors column.
    for row_index, _ in df_out.iterrows():
        total_error = 0
        for col_name in df_out.keys():
            if col_name not in NON_INPUT_DATA_COLUMNS:
                # value = df_out.at[row, col]
                new_value, error = closest_number(df_out.at[row_index, col_name])
                df_out.at[row_index, col_name] = new_value
                total_error += error
                # print(row, col, value, new_value, error, total_error)
        df_out.at[row_index, COL_ERRORS] = total_error

    return df_out


def recalculate_totals(df: pd.DataFrame) -> pd.DataFrame:
    """
    Recalculate the row and column totals. This funtion should be called
    after calling round_data so the totals are consistent with the data.
    :param df: input data frame
    :return: converted data frame
    """
    assert (isinstance(df, pd.DataFrame))

    # Make a copy of the data frame so the original is left untouched
    df_out = df.copy(deep=True)

    # Include the errors in the recalculations
    column_list = NON_INPUT_DATA_COLUMNS.copy()
    column_list.remove(COL_ERRORS)
    # print(column_list)

    # Calculate the row totals
    col_total = {col_name: 0.0 for col_name in df_out.keys()}
    # print(col_total)
    big_total = 0.0
    for row_index, row in df_out.iterrows():
        if row[COL_PROJECT_PATH] == ROW_TOTALS:  # it doesn't make sense to update this row yet
            continue
        row_total = 0.0
        for col_name in df_out.keys():
            # if col_name not in NON_INPUT_DATA_COLUMNS:
            if col_name not in column_list:
                # print(col_name)
                row_total += df_out.at[row_index, col_name]
                col_total[col_name] += df_out.at[row_index, col_name]
        big_total += row_total
        # print(row_total)

        df_out.at[row_index, COL_TOTALS] = row_total

    # Update column totals
    row = df_out.loc[df_out[COL_PROJECT_PATH] == ROW_TOTALS]
    row_index = row.index[0]
    # print(row)
    # print('--', row.index)
    for col_name in df_out.keys():
        # if col_name not in NON_INPUT_DATA_COLUMNS:
        if col_name not in column_list:
            df_out.at[row_index, col_name] = col_total[col_name]

    # Update total hours for period
    df_out.at[row_index, COL_TOTALS] = big_total

    return df_out


def convert_value(value: Any) -> Any:
    """
    Converts numpy.float64 values to float and replaces zero data with a '-'.
    All other data types are left unchanged.
    :param value: input value
    :return: converted value
    """
    # print('convert value', value, type(value))
    if isinstance(value, float64) or isinstance(value, int64):
        output_value = float(value)
        if abs(output_value) < 0.0001:
            output_value = '-'
        # print('found float64')
    else:
        output_value = value
    return output_value


def get_row_and_index(df: pd.DataFrame, name: str) -> tuple:
    """
    Auxiliary function that returns the row and the row index for the row where COL_PROJECT_PATH matches a given name.
    Provided to prevent code redundancy in the program.
    :param df: data frame
    :param name: name to match
    :return: row and row index
    """
    assert (isinstance(df, pd.DataFrame))
    row = df.loc[df[COL_PROJECT_PATH] == name]
    row_index = row.index[0]
    return row, row_index


def get_row_columns(df: pd.DataFrame, row_index: int, output_row_index: int, column_names: list) -> list:
    """
    Return the column values for a given row index. The output row index will be used
    when calculating the formulas since the order of rows in the output spreadsheet is
    not necessarily the same as the order of rows in the input data. The formula will
    be empty if the output
    :param df: data frame
    :param column_names: list of column names
    :param row_index: row index (zero indexed)
    :param output_row_index: output row index (used in formulas)
    :return: list of column values for the row
    """
    n = 0
    column_values = []
    for col_name in column_names:
        if col_name == COL_FORMULAS:
            if output_row_index is not None:
                value = '=sum(B' + str(output_row_index + 2) + ':' + excel_column_name(n - 2) + str(
                    output_row_index + 2) + ')'
            else:
                value = ''
        else:
            value = convert_value(df.at[row_index, col_name])
        column_values.append(value)
        n += 1
    return column_values


def excel_column_name(column_number: int) -> str:
    """
    Convert a zero-indexed column number into an Excel column name.
    :param column_number: column number (zero indexed)
    :return: column name
    """
    assert (isinstance(column_number, int))
    column_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I',
                   'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                   'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    max_columns = len(column_list)

    if column_number < max_columns:
        col_name = column_list[column_number]
    else:
        r = column_number // max_columns
        if r > max_columns:
            raise ValueError('column number too large')
        col_name = column_list[r - 1] + column_list[column_number - max_columns * r]

    return col_name


def write_excel_file(df: pd.DataFrame, project_list: list, file_name: str):
    """
    Write data to spreadsheet file
    :param df: data frame
    :param project_list: list of valid projects
    :param file_name: output file name
    """
    assert (isinstance(df, pd.DataFrame))

    # Create workbook and define basic attributes
    wb = Workbook()
    ws = wb.active
    ws.title = 'Timecard'

    # Get spreadsheet dimensions
    # Allocate two additional rows (column titles, and formulas)
    # Allocate an additional column for the formulas
    # row = df.loc[df[COL_PROJECT_PATH] == ROW_TOTALS]
    row, _ = get_row_and_index(df, ROW_TOTALS)
    n_rows = row.index[0] + 3  # row.index[0] is zero indexed
    n_cols = len(df.keys()) + 1
    # print(n_rows, n_cols)

    # Prepare the spreadsheet heading (column titles)
    column_titles = list(df.keys()).copy()
    column_titles = column_titles[:-1] + [COL_FORMULAS, COL_ERRORS]
    ws.append(column_titles)

    # Append the time data
    # Some formulas are inserted here
    project_row_index = 0
    for project_name in project_list:
        try:
            row, row_index = get_row_and_index(df, project_name)
        except IndexError:
            # print('ignoring', project_name, project_row_index)
            continue
        row_list = get_row_columns(df, row_index, project_row_index, column_titles)
        project_row_index += 1
        ws.append(row_list)

    # Append the row with the column totals
    row, row_index = get_row_and_index(df, ROW_TOTALS)
    row_list = get_row_columns(df, row_index, 0, column_titles)
    ws.append(row_list)

    # -- Formulas start here

    # Append the row containing the column totals (formulas)
    column_list = column_titles[1:-2]  # remove project name, totals and errors
    # print(column_list)
    row_list = [ROW_FORMULAS]
    for col_number in range(1, len(column_list) + 1):
        # print(col_number, excel_column_name(col_number))
        row_list.append('=sum(' + excel_column_name(col_number) + '2:' +
                        excel_column_name(col_number) + str(n_rows - 2) + ')')
    # Sum of all column totals
    row_list.append('=sum(B' + str(n_rows) + ':' +
                    excel_column_name(len(column_list) - 1) + str(n_rows) + ')')
    ws.append(row_list)

    # -- Font definitions start here

    # Change the font in all relevant cells
    ft = Font(name=FONT_NAME)
    for row_number in range(1, n_rows + 1):
        for col_number in range(0, n_cols):
            # print(excel_column_name(col_number) + str(row_number))
            ws[excel_column_name(col_number) + str(row_number)].font = ft

    # Format titles in bold
    ft = Font(name=FONT_NAME, bold=True)
    for col_number in range(0, n_cols):
        # print(excel_column_name(col_number) + '1')
        ws[excel_column_name(col_number) + '1'].font = ft

    # Format totals in bold. Use the same font definition used for the titles.
    for col_number in range(1, n_cols - 1):
        # print(excel_column_name(col_number) + str(n_rows - 1))
        ws[excel_column_name(col_number) + str(n_rows)].font = ft
    for row_number in range(1, n_rows):
        # print(excel_column_name(n_cols - 3) + str(row_number))
        ws[excel_column_name(n_cols - 3) + str(row_number)].font = ft

    # Change format and color in the formulas
    ft = Font(name=FONT_NAME, bold=True, italic=True, color="FF0000")
    for col_number in range(1, len(column_list) + 2):
        # print(excel_column_name(col_number))
        ws[excel_column_name(col_number) + str(n_rows)].font = ft
    for row_number in range(2, n_rows):
        # print(row_number)
        ws[excel_column_name(n_cols - 2) + str(row_number)].font = ft

    # -- Alignment and numeric format start here

    # Align numeric cells to the right
    for row_number in range(1, n_rows):
        for col_number in range(1, n_cols):
            # print(excel_column_name(col_number) + str(row_number))
            ws[excel_column_name(col_number) + str(row_number)].alignment = Alignment(horizontal='right')

    # Make numbers display with two decimals
    for row_number in range(1, n_rows + 1):
        for col_number in range(1, n_cols - 1):
            # print(excel_column_name(col_number) + str(row_number))
            ws[excel_column_name(col_number) + str(row_number)].number_format = '0.00'

    # Make errors display with four decimals
    for row_number in range(1, n_rows):
        # print(excel_column_name(n_cols - 1) + str(row_number))
        ws[excel_column_name(n_cols - 1) + str(row_number)].number_format = '0.0000'

    # Save output file
    wb.save(file_name)


def get_args(argv: list) -> Namespace:
    """
    Process command line arguments
    :param argv: command line arguments from sys.argv
    :return: arguments
    """

    parser = ArgumentParser()

    parser.add_argument(action='store',
                        dest='file',
                        default='')

    parser.add_argument('--debug',
                        action='store_true',
                        dest='debug',
                        default=False,
                        help=SUPPRESS)

    return parser.parse_args(argv[1:])


def read_csv_file(file_name):
    df = pd.read_csv(file_name)

    # Drop unused columns
    df_out = df.drop(columns=UNNEEDED_COLUMNS)

    # Look for the row containing the column totals
    lst = df_out.index[df_out[COL_PROJECT_PATH] == ROW_TOTALS].tolist()
    if len(lst) == 1:
        index = lst[0]
    else:
        raise (ValueError, 'No totals')

    # print(df_out[0:index + 1])
    # return

    # Only return rows up to the totals
    return df_out[0:index + 1]


if __name__ == '__main__':

    # Get command line arguments
    args = get_args(sys.argv)
    # args = get_args(['ptimecard', 'example.csv', '--debug'])
    input_file = args.file
    output_file = 'Output_' + str(input_file).replace('.csv', '') + '.xlsx'
    debug = args.debug

    master_project_list = []  # to make pycharm happy
    try:
        master_project_list = read_project_list(PROJECT_FILE)
    except FileNotFoundError as e:
        print('Project file not found', e)
        exit(0)
    if debug:
        print(master_project_list)

    data_frame_excel = None  # to make pycharm happy
    try:
        data_frame_excel = read_csv_file(input_file)
    except Exception as e:
        print('Cannot open spreadsheet', e)
        exit(0)

    if debug:
        dump_data(data_frame_excel, label='read_excel', print_types=True)

    if not check_projects(data_frame_excel, master_project_list):
        exit(0)

    data_frame_convert = convert_values(data_frame_excel)
    if debug:
        dump_data(data_frame_convert, 'convert_values', print_types=True)

    data_frame_rename = rename_columns(data_frame_convert)
    if debug:
        dump_data(data_frame_rename, label='rename_columns')

    data_frame_round = round_data(data_frame_rename)
    if debug:
        dump_data(data_frame_round, label='round_data')

    data_frame_recalculate = recalculate_totals(data_frame_round)
    if debug:
        dump_data(data_frame_recalculate, label='recalculate_totals')

    try:
        write_excel_file(data_frame_recalculate, master_project_list, output_file)
    except Exception as e:
        print('Cannot write spreadsheet', e)
