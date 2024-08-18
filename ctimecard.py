"""
Convert a clockify monthly detailed report into a time card
for the first or second part of the month.
"""
import sys
import pandas as pd
from argparse import ArgumentParser, Namespace
from openpyxl import Workbook
from openpyxl.styles import Alignment
# from openpyxl.styles import Font, Alignment
from timecard import Totals, read_project_list, column_name

# File containing the valid project names
PROJECT_FILE = 'clockify_projects.txt'

# Input column names
COL_PROJECT = 'Project'
COL_CLIENT = 'Client'
COL_DESCRIPTION = 'Description'
COL_TASK = 'Task'
COL_USER = 'User'
COL_GROUP = 'Group'
COL_EMAIL = 'Email'
COL_TAGS = 'Tags'
COL_BILLABLE = 'Billable'
COL_START_DATE = 'Start Date'
COL_START_TIME = 'Start Time'
COL_END_DATE = 'End Date'
COL_END_TIME = 'End Time'
COL_DURATION = 'Duration (h)'
COL_DURATION_DEC = 'Duration (decimal)'
COL_BILLABLE_RATE = 'Billable Rate (USD)'
COL_BILLABLE_AMOUNT = 'Billable Amount (USD)'

# Output column names
COL_TOTALS = 'Totals'
COL_ERRORS = 'Errors'
COL_FORMULAS = 'Formulas'

# Output row names
ROW_TOTALS = 'Totals'
ROW_FORMULAS = 'Formulas'

# Column types
UNNEEDED_COLUMNS = [COL_CLIENT, COL_DESCRIPTION, COL_TASK, COL_USER, COL_GROUP,
                    COL_EMAIL, COL_TAGS, COL_BILLABLE, COL_START_TIME, COL_END_TIME,
                    COL_DURATION_DEC, COL_BILLABLE_RATE, COL_BILLABLE_AMOUNT]

EMPTY_CELL = '-'


def read_csv_file(file_name) -> pd.DataFrame:
    """
    Read a csv file containing time tracking information into a dataframe
    :param file_name: input file name
    :return: data frame
    """
    df_in = pd.read_csv(file_name)
    df_out = df_in.drop(columns=UNNEEDED_COLUMNS)
    return df_out


def process_data(data: pd.DataFrame, project_list: list, start_day: int, end_day: int) -> Totals:
    """
    Process the data in contained in a data frame. It returns a Totals object
    containing the processed data.
    :param data: time data
    :param project_list: project list
    :param start_day: starting day
    :param end_day: ending day
    :return: Total object
    """
    output = Totals()
    for index, row in data.iterrows():
        project, start_date, end_date, hours = (row[COL_PROJECT], row[COL_START_DATE],
                                                row[COL_END_DATE], row[COL_DURATION])
        # print(index, project, start_date, end_date, hours)
        if project in project_list:
            if start_date == end_date:
                output.add(project, start_date, hours, start_day, end_day)
            else:
                raise ValueError(f'date mismatch for entry ({project}) {start_date}, {end_date}')
        else:
            raise ValueError(f'project "{project}" not found')
    return output


def write_timesheet(time_data: Totals, project_list: list,
                    start_day: int, end_day: int, file_name: str):
    """
    Write data into an Excel spreadsheet
    :param time_data:
    :param project_list:
    :param start_day:
    :param end_day:
    :param file_name:
    :return:
    """
    # Create workbook and define basic attributes
    wb = Workbook()
    ws = wb.active
    ws.title = 'Timecard'

    # Get sheet dimensions
    # Allocate two additional rows (column titles, totals, and formulas)
    # Allocate three additional columns (totals, formulas and errors)
    n_rows = len(time_data.get_project_list()) + 3

    day_list = list(range(start_day, end_day + 1))
    len_day_list = len(day_list)

    # Prepare the spreadsheet heading (column titles)
    column_titles = [COL_PROJECT] + day_list + [COL_TOTALS, COL_ERRORS, COL_FORMULAS]
    ws.append(column_titles)

    # Write rows with daily data
    row = 2
    for project_name in project_list:
        row_list = [project_name]
        for day in day_list:
            row_list.append(totals.get_hours(project_name, day))
        hours, error = totals.get_project_totals(project_name)
        row_list.extend([hours, error])
        row_list.append('=sum(B' + str(row) + ':' +
                        column_name(len_day_list) + str(row) + ')')
        row += 1
        ws.append(row_list)

    # Write daily totals
    row_list = ['Totals']
    for day in day_list:
        row_list.append(totals.get_day_totals(day))
    row_list.append(totals.get_total_time())
    ws.append(row_list)

    # Write column with formulas
    row_list = [ROW_FORMULAS]
    for column in range(1, len(day_list) + 1):
        row_list.append('=sum(' + column_name(column) + '2:' +
                        column_name(column) + str(n_rows - 2) + ')')
    row_list.append('')
    row_list.append('=sum(' + column_name(len_day_list + 2) + '2:' +
                    column_name(len_day_list + 2) + str(n_rows - 2) + ')')
    row_list.append('=sum(' + column_name(len_day_list + 3) + '2:' +
                    column_name(len_day_list + 3) + str(n_rows - 2) + ')')
    ws.append(row_list)

    # Replace data cells containing a zero value with a special marker
    # Skip the rows and columns containing totals, errors or formulas
    for row in range(1, ws.max_row - 1):
        for col in range(1, ws.max_column - 2):
            # print(row, col, ws.cell(row, col).value, type(ws.cell(row, col).value))
            try:
                if float(ws.cell(row, col).value) < 0.000001:
                    ws.cell(row, col).value = EMPTY_CELL
                    ws.cell(row, col).alignment = Alignment(horizontal='right')
            except (ValueError, TypeError):
                pass

    # Save output file
    wb.save(file_name)


def get_args(argv: list) -> Namespace:
    """
    Process command line arguments
    :param argv: command line arguments from sys.argv
    :return: arguments
    """

    parser = ArgumentParser()

    parser.add_argument(action='store',
                        dest='file',
                        help='input csv file',
                        default='')

    parser.add_argument('-s', '--start',
                        action='store',
                        dest='start_day',
                        help='Starting day',
                        default=1)

    parser.add_argument('-e', '--end',
                        action='store',
                        dest='end_day',
                        help='Ending day',
                        default=31)

    return parser.parse_args(argv[1:])


if __name__ == '__main__':

    # Get command line arguments
    args = get_args(sys.argv)
    # args = get_args(['program', 'Timesheet20240815.csv', '-s', '1', '-e', '31'])
    input_file = str(args.file)
    output_file = 'Output_' + input_file.replace('.csv', '') + '.xlsx'
    try:
        starting_day = int(args.start_day)
        ending_day = int(args.end_day)
    except ValueError:
        print('Bad starting or ending day')
        exit(0)
    print(f'Day range: [{starting_day}, {ending_day}]')

    # Read project list
    try:
        master_project_list = read_project_list(PROJECT_FILE)
    except FileNotFoundError as e:
        print('Project file not found', e)
        exit(0)

    # Read clockify data into a pandas dataframe
    try:
        df = read_csv_file(input_file)
    except ValueError as e:
        print(e)
        exit(0)

    # Process the data. The total time per day and project
    # is calculated at this point.
    totals = process_data(df, master_project_list, starting_day, ending_day)

    # Get the list of projects in the data in the right order
    # aux = totals.get_project_list()
    sorted_project_list = [_ for _ in master_project_list if _ in totals.get_project_list()]
    # print(sorted_project_list)

    # totals.dump()
    # print('project list', totals.get_project_list())
    # print('totals project', totals.get_project_totals())
    # print('totals date', totals.get_date_totals())

    # print(totals.get_number_of_days())
    if len(totals) > 0:
        write_timesheet(totals, sorted_project_list,
                        starting_day, ending_day, output_file)
    else:
        print('no data for the day range')
